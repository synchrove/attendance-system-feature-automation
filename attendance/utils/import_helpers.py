# Python standard library imports
import csv
import io
import os
from datetime import datetime, date, timedelta

# Local model imports
from ..models import AttendanceRecord, Employee

# Optional Excel support with graceful fallback
try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


def read_import_file(upload_file):
    """
    Universal file reader supporting CSV and Excel formats.
    
    Supported formats:
    - CSV files (UTF-8 with BOM support)
    - Excel files (.xlsx, .xlsm) if openpyxl available
    - Automatic format detection based on filename and content type
    
    Process:
    1. Detect file format from filename/content type
    2. Parse headers from first row
    3. Extract data rows as lists
    4. Handle encoding issues gracefully
    
    Returns: Tuple of (headers_list, rows_list)
    """
    headers = []
    rows_iter = []
    filename = getattr(upload_file, "name", "upload")
    content_type = upload_file.content_type

    # Try Excel first (if openpyxl available and looks like spreadsheet)
    if HAS_OPENPYXL and (
        filename.lower().endswith((".xlsx", ".xlsm")) or "spreadsheet" in content_type
    ):
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(upload_file.read()), data_only=True
        )
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c is not None else "" for c in row]
                continue
            rows_iter.append(list(row))
    else:
        text = upload_file.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        for i, row in enumerate(reader):
            if i == 0:
                headers = [c.strip() for c in row]
                continue
            rows_iter.append(row)

    return headers, rows_iter


def build_header_mapping(headers):
    """
    Intelligent header mapping with fuzzy matching.
    
    Supported field mappings:
    - employee_id: 'employee id', 'employee_id', variations
    - date: 'date'
    - checkin_time/checkout_time: Various time field names
    - status: 'status'
    - shift_name: 'shift name', 'shift_name'
    - image files: checkin/checkout image file paths
    
    Features:
    - Case-insensitive matching
    - Flexible field name recognition
    - Handles spaces and underscores
    
    Returns: Dictionary mapping logical_field -> column_index
    """
    headers_norm = [h.lower().strip() for h in headers]
    mapping = {}

    for idx, h in enumerate(headers_norm):
        if "employee" in h and "id" in h:
            mapping["employee_id"] = idx
        elif h == "employee_id":
            mapping["employee_id"] = idx
        elif h == "date":
            mapping["date"] = idx
        elif "checkin" in h and "time" in h:
            mapping["checkin_time"] = idx
        elif "checkout" in h and "time" in h:
            mapping["checkout_time"] = idx
        elif h == "checkin_time":
            mapping["checkin_time"] = idx
        elif h == "checkout_time":
            mapping["checkout_time"] = idx
        elif "status" in h:
            mapping["status"] = idx
        elif "late" in h and "second" in h:
            mapping["late_seconds"] = idx
        elif "checkin" in h and "image" in h:
            mapping["checkin_image_file"] = idx
        elif "checkout" in h and "image" in h:
            mapping["checkout_image_file"] = idx
        elif "shift" in h and "name" in h:
            mapping["shift_name"] = idx
        elif h == "shift_name":
            mapping["shift_name"] = idx

    return mapping


def parse_any_date(raw_date_str, raw_cell):
    """
    Robust date parsing supporting multiple formats.
    
    Supported formats:
    - ISO format: YYYY-MM-DD (preferred)
    - European: DD/MM/YYYY
    - American: MM/DD/YYYY
    - Excel numeric dates (days since 1899-12-30)
    
    Features:
    - Tries formats in order of reliability
    - Handles Excel's numeric date representation
    - Graceful failure with None return
    
    Returns: date object or None if parsing fails
    """
    if not raw_date_str:
        return None

    # Try ISO
    try:
        return date.fromisoformat(raw_date_str)
    except Exception:
        pass

    # dd/mm/yyyy (prioritize this format)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw_date_str, fmt).date()
        except Exception:
            continue

    # Excel numeric date as fallback
    if isinstance(raw_cell, (int, float)):
        try:
            return date(1899, 12, 30) + timedelta(days=float(raw_cell))
        except Exception:
            pass

    return None


def parse_any_time(raw_value, parsed_date):
    """
    Advanced time parsing with timezone awareness.
    
    Supported formats:
    - 12-hour with AM/PM: '07:55 AM', '05:30 PM'
    - 12-hour with seconds: '07:55:30 AM'
    - 24-hour format: '07:55', '17:30:00'
    - Full ISO datetime strings
    
    Features:
    - Automatic timezone localization to Asia/Dhaka
    - Combines time with provided date
    - Handles both naive and timezone-aware inputs
    - Critical for mobile app integration
    
    Returns: timezone-aware datetime object or None
    """
    if raw_value in (None, ""):
        return None

    raw_str = str(raw_value).strip()
    
    # Import timezone here to match model
    import pytz
    dhaka = pytz.timezone("Asia/Dhaka")

    # Full ISO datetime
    try:
        dt = datetime.fromisoformat(raw_str)
        if dt.tzinfo:
            return dt.astimezone(dhaka)
        else:
            return dhaka.localize(dt)
    except Exception:
        pass

    # Time formats with AM/PM and 24-hour
    time_formats = [
        "%I:%M:%S %p",  # 12-hour with seconds and AM/PM
        "%I:%M %p",     # 12-hour with AM/PM
        "%H:%M:%S",     # 24-hour with seconds
        "%H:%M",        # 24-hour
    ]
    
    for fmt in time_formats:
        try:
            t = datetime.strptime(raw_str, fmt).time()
            dt = datetime.combine(parsed_date, t)
            # Localize to Dhaka timezone to match model
            return dhaka.localize(dt)
        except Exception:
            continue

    return None


def handle_export(request, selected_year, selected_month):
    """
    Comprehensive export system with image packaging.
    
    Export format:
    - ZIP file containing Excel data + organized images
    - Excel file: attendance_data.xlsx with all fields
    - Images: organized in checkin/ and checkout/ folders
    - Filename format: attendance_YYYY_MM.zip
    
    Features:
    - Timezone-aware time formatting (12-hour AM/PM)
    - Image path tracking in Excel
    - Organized folder structure
    - Complete data preservation for re-import
    
    Returns: HttpResponse with ZIP file or None if no data
    """
    if request.method != "POST" or not request.POST.get("export_data"):
        return None
    
    import zipfile
    import os
    from django.http import HttpResponse
    from django.conf import settings
    
    # Get attendance records for selected month
    records = AttendanceRecord.objects.filter(
        date__year=selected_year,
        date__month=selected_month
    ).select_related('employee', 'shift').order_by('date', 'employee__employee_id')
    
    if not records.exists():
        return None
    
    # Create ZIP response
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="attendance_{selected_year}_{selected_month:02d}.zip"'
    
    with zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Create Excel data
        if HAS_OPENPYXL:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Data"
            
            # Headers
            headers = [
                'employee_id', 'name', 'department', 'designation', 'date',
                'checkin_time', 'checkout_time', 'status', 'late_duration_seconds',
                'device_id', 'shift_name', 'checkin_image_file', 'checkout_image_file'
            ]
            ws.append(headers)
            
            # Data rows
            for record in records:
                checkin_file = ''
                checkout_file = ''
                
                # Add images to ZIP
                if record.checkin_image:
                    try:
                        img_path = record.checkin_image.path
                        if os.path.exists(img_path):
                            checkin_file = f"images/checkin/{record.employee.employee_id}_{record.date}_in.jpg"
                            zip_file.write(img_path, checkin_file)
                    except:
                        pass
                
                if record.checkout_image:
                    try:
                        img_path = record.checkout_image.path
                        if os.path.exists(img_path):
                            checkout_file = f"images/checkout/{record.employee.employee_id}_{record.date}_out.jpg"
                            zip_file.write(img_path, checkout_file)
                    except:
                        pass
                
                # Convert times to Dhaka timezone for proper display
                # Critical: Export times in same format as dashboard (12-hour AM/PM)
                import pytz
                dhaka = pytz.timezone("Asia/Dhaka")
                
                checkin_display = ''
                checkout_display = ''
                
                if record.checkin_time:
                    checkin_dhaka = record.checkin_time.astimezone(dhaka)
                    checkin_display = checkin_dhaka.strftime('%I:%M %p')
                
                if record.checkout_time:
                    checkout_dhaka = record.checkout_time.astimezone(dhaka)
                    checkout_display = checkout_dhaka.strftime('%I:%M %p')
                
                row = [
                    record.employee.employee_id,
                    record.employee.name,
                    record.employee.department,
                    record.employee.designation,
                    record.date.strftime('%d/%m/%Y'),
                    checkin_display,
                    checkout_display,
                    record.status or '',
                    int(record.late_duration.total_seconds()) if record.late_duration else 0,
                    record.device_id or '',
                    record.shift.name if record.shift else '',
                    checkin_file,
                    checkout_file
                ]
                ws.append(row)
            
            # Save Excel to ZIP
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr('attendance_data.xlsx', excel_buffer.getvalue())
    
    return response


def handle_zip_import(zip_file, selected_year, selected_month):
    """
    Advanced ZIP import with image restoration.
    
    Process:
    1. Extract ZIP to temporary directory
    2. Locate Excel file (attendance_data.xlsx)
    3. Parse attendance data with header mapping
    4. Import times, status, and shift information
    5. Restore images from organized folders
    6. Create/update attendance records
    7. Clean up temporary files
    
    Features:
    - Complete data restoration from export
    - Image file restoration with proper naming
    - Shift information preservation
    - Error handling with detailed reporting
    
    Returns: Tuple of (error_list, success_stats)
    """
    import zipfile
    import tempfile
    import shutil
    from django.core.files import File
    from django.conf import settings
    
    import_errors = []
    created = 0
    updated = 0
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Extract ZIP
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Find Excel file
        excel_file = None
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.endswith('.xlsx'):
                    excel_file = os.path.join(root, file)
                    break
        
        if not excel_file:
            import_errors.append("No Excel file found in ZIP")
            return import_errors, {"created": 0, "updated": 0}
        
        # Read Excel data
        if HAS_OPENPYXL:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            mapping = build_header_mapping(headers)
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    emp_id = str(row[mapping["employee_id"]]).strip() if mapping.get("employee_id") is not None else ""
                    raw_date = str(row[mapping["date"]]).strip() if mapping.get("date") is not None else ""
                    
                    if not emp_id or not raw_date:
                        continue
                    
                    parsed_date = parse_any_date(raw_date, row[mapping["date"]])
                    if not parsed_date or parsed_date.year != selected_year or parsed_date.month != selected_month:
                        continue
                    
                    try:
                        emp = Employee.objects.get(employee_id=emp_id)
                    except Employee.DoesNotExist:
                        import_errors.append(f"Row {row_idx}: employee '{emp_id}' not found")
                        continue
                    
                    obj, created_flag = AttendanceRecord.objects.get_or_create(employee=emp, date=parsed_date)
                    changed = False
                    
                    # Import times and status
                    ci = None
                    co = None
                    
                    if mapping.get("checkin_time") is not None and row[mapping["checkin_time"]]:
                        ci = parse_any_time(row[mapping["checkin_time"]], parsed_date)
                        if ci and obj.checkin_time != ci:
                            obj.checkin_time = ci
                            changed = True
                    
                    if mapping.get("checkout_time") is not None and row[mapping["checkout_time"]]:
                        co = parse_any_time(row[mapping["checkout_time"]], parsed_date)
                        if co and obj.checkout_time != co:
                            obj.checkout_time = co
                            changed = True
                    
                    if mapping.get("status"):
                        status = str(row[mapping["status"]]).strip() if row[mapping["status"]] else ""
                        if status and obj.status != status:
                            obj.status = status
                            changed = True
                    
                    # Import shift if available
                    if mapping.get("shift_name") is not None and row[mapping["shift_name"]]:
                        shift_name = str(row[mapping["shift_name"]]).strip()
                        if shift_name:
                            from ..models import Shift
                            try:
                                shift = Shift.objects.get(name=shift_name)
                                if obj.shift != shift:
                                    obj.shift = shift
                                    changed = True
                            except Shift.DoesNotExist:
                                import_errors.append(f"Row {row_idx}: shift '{shift_name}' not found")
                    
                    # Import images
                    if mapping.get("checkin_image_file"):
                        img_path = str(row[mapping["checkin_image_file"]]).strip() if row[mapping["checkin_image_file"]] else ""
                        if img_path:
                            full_img_path = os.path.join(temp_dir, img_path)
                            if os.path.exists(full_img_path):
                                with open(full_img_path, 'rb') as img_file:
                                    obj.checkin_image.save(f"{emp_id}_{parsed_date}_in.jpg", File(img_file), save=False)
                                    changed = True
                    
                    if mapping.get("checkout_image_file"):
                        img_path = str(row[mapping["checkout_image_file"]]).strip() if row[mapping["checkout_image_file"]] else ""
                        if img_path:
                            full_img_path = os.path.join(temp_dir, img_path)
                            if os.path.exists(full_img_path):
                                with open(full_img_path, 'rb') as img_file:
                                    obj.checkout_image.save(f"{emp_id}_{parsed_date}_out.jpg", File(img_file), save=False)
                                    changed = True
                    
                    if created_flag:
                        obj.save()
                        created += 1
                    elif changed:
                        obj.save()
                        updated += 1
                        
                except Exception as e:
                    import_errors.append(f"Row {row_idx}: {str(e)}")
    
    return import_errors, {"created": created, "updated": updated}


def handle_import(request, selected_year, selected_month):
    """
    Universal import handler supporting multiple formats.
    
    Supported formats:
    - CSV files (.csv)
    - Excel files (.xlsx, .xlsm)
    - ZIP packages (.zip) with images
    
    Process:
    1. Detect file format from extension
    2. Route to appropriate handler (CSV/Excel/ZIP)
    3. Validate required columns (employee_id, date)
    4. Parse and import data with error tracking
    5. Update existing records or create new ones
    6. Preserve shift information and images
    
    Features:
    - Month/year filtering (only imports selected period)
    - Employee validation (must exist in system)
    - Comprehensive error reporting
    - Statistics tracking (created/updated counts)
    
    Returns: Tuple of (error_messages_list, success_statistics_dict)
    """
    import_errors = []
    import_success = None

    if request.method != "POST" or not request.FILES.get("import_file"):
        return import_errors, import_success

    f = request.FILES["import_file"]
    
    # Check if it's a ZIP file
    if f.name.lower().endswith('.zip'):
        return handle_zip_import(f, selected_year, selected_month)

    try:
        headers, rows_iter = read_import_file(f)
    except Exception as e:
        import_errors.append(f"Failed to read file: {e}")
        return import_errors, import_success

    mapping = build_header_mapping(headers)

    # require employee_id and date
    if "employee_id" not in mapping or "date" not in mapping:
        import_errors.append("File must include at least columns: employee_id, date")
        return import_errors, import_success

    created = 0
    updated = 0

    for ridx, row in enumerate(rows_iter, start=2):
        try:
            
            emp_id = (
                str(row[mapping["employee_id"]]).strip()
                if mapping.get("employee_id") is not None and len(row) > mapping["employee_id"]
                else ""
            )
            raw_date = (
                str(row[mapping["date"]]).strip()
                if mapping.get("date") is not None and len(row) > mapping["date"]
                else ""
            )

            if not emp_id or not raw_date:
                import_errors.append(
                    f"Row {ridx}: missing employee_id '{emp_id}' or date '{raw_date}'; skipping"
                )
                continue

            parsed_date = parse_any_date(raw_date, row[mapping["date"]] if len(row) > mapping["date"] else None)
            if not parsed_date:
                import_errors.append(f"Row {ridx}: unrecognized date '{raw_date}'")
                continue

            # Only import selected month/year
            if parsed_date.year != selected_year or parsed_date.month != selected_month:
                continue

            # Employee lookup
            try:
                emp = Employee.objects.get(employee_id=emp_id)
            except Employee.DoesNotExist:
                import_errors.append(
                    f"Row {ridx}: employee_id '{emp_id}' not found; skipping"
                )
                continue

            obj, created_flag = AttendanceRecord.objects.get_or_create(
                employee=emp, date=parsed_date
            )
            changed = False

            # parse times only if columns exist and have data
            if mapping.get("checkin_time") is not None and len(row) > mapping["checkin_time"] and row[mapping["checkin_time"]]:
                raw_checkin = row[mapping["checkin_time"]]
                ci = parse_any_time(raw_checkin, parsed_date)

                if ci and obj.checkin_time != ci:
                    obj.checkin_time = ci
                    changed = True

            
            if mapping.get("checkout_time") is not None and len(row) > mapping["checkout_time"] and row[mapping["checkout_time"]]:
                raw_checkout = row[mapping["checkout_time"]]
                co = parse_any_time(raw_checkout, parsed_date)

                if co and obj.checkout_time != co:
                    obj.checkout_time = co
                    changed = True


            if mapping.get("status") is not None and len(row) > mapping["status"] and row[mapping["status"]]:
                val = row[mapping["status"]]
                if val not in (None, ""):
                    s = str(val).strip()
                    if s and obj.status != s:
                        obj.status = s
                        changed = True
            
            # Import shift if available
            if mapping.get("shift_name") is not None and len(row) > mapping["shift_name"] and row[mapping["shift_name"]]:
                shift_name = str(row[mapping["shift_name"]]).strip()
                if shift_name:
                    from ..models import Shift
                    try:
                        shift = Shift.objects.get(name=shift_name)
                        if obj.shift != shift:
                            obj.shift = shift
                            changed = True
                    except Shift.DoesNotExist:
                        import_errors.append(f"Row {ridx}: shift '{shift_name}' not found")
            
            if created_flag:
                obj.save()
                created += 1

            elif changed:
                obj.save()
                updated += 1


        except Exception as e:
            import_errors.append(f"Row {ridx}: unexpected error: {e}")

    import_success = {"created": created, "updated": updated}
    return import_errors, import_success


def export_employees(request):
    """
    Export employee data as ZIP with Excel and images.
    
    Features:
    - Complete employee information export
    - Employee photos included in ZIP
    - Department/designation filtering support
    - Excel format with proper headers
    
    Returns: HttpResponse with ZIP file
    """
    from django.http import HttpResponse
    import io
    
    # Apply same filters as view
    selected_department = request.GET.get('department') or None
    selected_designation = request.GET.get('designation') or None
    search_query = request.GET.get('search', '').strip()
    
    employees_qs = Employee.objects.all()
    
    if selected_department:
        employees_qs = employees_qs.filter(department=selected_department)
    if selected_designation:
        employees_qs = employees_qs.filter(designation=selected_designation)
    if search_query:
        from django.db.models import Q
        employees_qs = employees_qs.filter(
            Q(name__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    employees_qs = employees_qs.order_by('employee_id')
    
    # Create ZIP response
    import zipfile
    import os
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="employees.zip"'
    
    with zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        if HAS_OPENPYXL:
            # Excel export
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employees"
            
            # Headers
            headers = [
                'employee_id', 'name', 'email', 'phone', 'department', 
                'designation', 'monthly_salary', 'hire_date',
                'employee_image_file', 'face_encoding_status'
            ]
            ws.append(headers)
            
            # Data rows
            for emp in employees_qs:
                image_file = ''
                if emp.employee_image:
                    try:
                        img_path = emp.employee_image.path
                        if os.path.exists(img_path):
                            image_file = f"images/{emp.employee_id}.jpg"
                            zip_file.write(img_path, image_file)
                    except:
                        pass
                
                face_status = 'Yes' if emp.face_encoding else 'No'
                
                row = [
                    emp.employee_id,
                    emp.name,
                    emp.email or '',
                    emp.phone or '',
                    emp.department or '',
                    emp.designation or '',
                    float(emp.monthly_salary) if emp.monthly_salary else 0,
                    emp.hire_date.strftime('%d/%m/%Y') if emp.hire_date else '',
                    image_file,
                    face_status
                ]
                ws.append(row)
            
            # Save Excel to ZIP
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr('employees.xlsx', excel_buffer.getvalue())
    
    return response


def import_employees(request):
    """
    Import employee data from CSV/Excel/ZIP files.
    
    Supported fields:
    - employee_id (required)
    - name (required)
    - email, phone, department, designation
    - monthly_salary, hire_date
    - employee_image_file (for ZIP imports)
    
    Features:
    - Creates new employees or updates existing ones
    - Validates required fields
    - Handles date parsing for hire_date
    - Salary parsing with decimal support
    - Image import from ZIP files
    
    Returns: Tuple of (error_messages_list, success_statistics_dict)
    """
    import_errors = []
    import_success = None
    
    f = request.FILES.get('import_file')
    if not f:
        return import_errors, import_success
    
    # Check if it's a ZIP file
    if f.name.lower().endswith('.zip'):
        return import_employees_zip(f)
    
    try:
        headers, rows_iter = read_import_file(f)
    except Exception as e:
        import_errors.append(f"Failed to read file: {e}")
        return import_errors, import_success
    
    # Build header mapping for employee fields
    headers_norm = [h.lower().strip() for h in headers]
    mapping = {}
    
    for idx, h in enumerate(headers_norm):
        if "employee" in h and "id" in h:
            mapping["employee_id"] = idx
        elif h == "employee_id":
            mapping["employee_id"] = idx
        elif h == "name":
            mapping["name"] = idx
        elif h == "email":
            mapping["email"] = idx
        elif h == "phone":
            mapping["phone"] = idx
        elif h == "department":
            mapping["department"] = idx
        elif h == "designation":
            mapping["designation"] = idx
        elif "salary" in h:
            mapping["monthly_salary"] = idx
        elif "hire" in h and "date" in h:
            mapping["hire_date"] = idx
        elif h == "hire_date":
            mapping["hire_date"] = idx
        elif "active" in h:
            mapping["is_active"] = idx
    
    # Require employee_id and name
    if "employee_id" not in mapping or "name" not in mapping:
        import_errors.append("File must include at least columns: employee_id, name")
        return import_errors, import_success
    
    created = 0
    updated = 0
    
    for ridx, row in enumerate(rows_iter, start=2):
        try:
            emp_id = (
                str(row[mapping["employee_id"]]).strip()
                if len(row) > mapping["employee_id"] and row[mapping["employee_id"]]
                else ""
            )
            name = (
                str(row[mapping["name"]]).strip()
                if len(row) > mapping["name"] and row[mapping["name"]]
                else ""
            )
            
            if not emp_id or not name:
                import_errors.append(
                    f"Row {ridx}: missing employee_id '{emp_id}' or name '{name}'; skipping"
                )
                continue
            
            # Get or create employee
            emp, created_flag = Employee.objects.get_or_create(
                employee_id=emp_id,
                defaults={'name': name}
            )
            changed = False
            
            # Update name if different
            if emp.name != name:
                emp.name = name
                changed = True
            
            # Update other fields if present
            if mapping.get("email") is not None and len(row) > mapping["email"] and row[mapping["email"]]:
                email = str(row[mapping["email"]]).strip()
                if email and emp.email != email:
                    emp.email = email
                    changed = True
            
            if mapping.get("phone") is not None and len(row) > mapping["phone"] and row[mapping["phone"]]:
                phone = str(row[mapping["phone"]]).strip()
                if phone and emp.phone != phone:
                    emp.phone = phone
                    changed = True
            
            if mapping.get("department") is not None and len(row) > mapping["department"] and row[mapping["department"]]:
                dept = str(row[mapping["department"]]).strip()
                if dept and emp.department != dept:
                    emp.department = dept
                    changed = True
            
            if mapping.get("designation") is not None and len(row) > mapping["designation"] and row[mapping["designation"]]:
                desig = str(row[mapping["designation"]]).strip()
                if desig and emp.designation != desig:
                    emp.designation = desig
                    changed = True
            
            if mapping.get("monthly_salary") is not None and len(row) > mapping["monthly_salary"] and row[mapping["monthly_salary"]]:
                try:
                    salary = float(str(row[mapping["monthly_salary"]]).strip())
                    if salary and emp.monthly_salary != salary:
                        emp.monthly_salary = salary
                        changed = True
                except ValueError:
                    import_errors.append(f"Row {ridx}: invalid salary value")
            
            if mapping.get("hire_date") is not None and len(row) > mapping["hire_date"] and row[mapping["hire_date"]]:
                hire_date = parse_any_date(
                    str(row[mapping["hire_date"]]).strip(),
                    row[mapping["hire_date"]]
                )
                if hire_date and emp.hire_date != hire_date:
                    emp.hire_date = hire_date
                    changed = True
            
            if mapping.get("is_active") is not None and len(row) > mapping["is_active"] and row[mapping["is_active"]]:
                active_str = str(row[mapping["is_active"]]).strip().lower()
                is_active = active_str in ('yes', 'true', '1', 'active')
                if emp.is_active != is_active:
                    emp.is_active = is_active
                    changed = True
            
            if created_flag:
                emp.save()
                created += 1
            elif changed:
                emp.save()
                updated += 1
                
        except Exception as e:
            import_errors.append(f"Row {ridx}: unexpected error: {e}")
    
    import_success = {"created": created, "updated": updated}
    return import_errors, import_success

def import_employees_zip(zip_file):
    """
    Import employees from ZIP file with images.
    
    Process:
    1. Extract ZIP to temporary directory
    2. Locate Excel file (employees.xlsx)
    3. Parse employee data with header mapping
    4. Import employee photos from images folder
    5. Create/update employee records
    6. Clean up temporary files
    
    Returns: Tuple of (error_list, success_stats)
    """
    import zipfile
    import tempfile
    import shutil
    from django.core.files import File
    from django.conf import settings
    
    import_errors = []
    created = 0
    updated = 0
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Extract ZIP
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Find Excel file
        excel_file = None
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.endswith('.xlsx'):
                    excel_file = os.path.join(root, file)
                    break
        
        if not excel_file:
            import_errors.append("No Excel file found in ZIP")
            return import_errors, {"created": 0, "updated": 0}
        
        # Read Excel data
        if HAS_OPENPYXL:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            
            # Build header mapping for employee fields
            headers_norm = [h.lower().strip() for h in headers]
            mapping = {}
            
            for idx, h in enumerate(headers_norm):
                if "employee" in h and "id" in h:
                    mapping["employee_id"] = idx
                elif h == "employee_id":
                    mapping["employee_id"] = idx
                elif h == "name":
                    mapping["name"] = idx
                elif h == "email":
                    mapping["email"] = idx
                elif h == "phone":
                    mapping["phone"] = idx
                elif h == "department":
                    mapping["department"] = idx
                elif h == "designation":
                    mapping["designation"] = idx
                elif "salary" in h:
                    mapping["monthly_salary"] = idx
                elif "hire" in h and "date" in h:
                    mapping["hire_date"] = idx
                elif h == "hire_date":
                    mapping["hire_date"] = idx
                elif "image" in h and "file" in h:
                    mapping["employee_image_file"] = idx
            
            # Require employee_id and name
            if "employee_id" not in mapping or "name" not in mapping:
                import_errors.append("File must include at least columns: employee_id, name")
                return import_errors, {"created": 0, "updated": 0}
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    emp_id = (
                        str(row[mapping["employee_id"]]).strip()
                        if len(row) > mapping["employee_id"] and row[mapping["employee_id"]]
                        else ""
                    )
                    name = (
                        str(row[mapping["name"]]).strip()
                        if len(row) > mapping["name"] and row[mapping["name"]]
                        else ""
                    )
                    
                    if not emp_id or not name:
                        import_errors.append(
                            f"Row {row_idx}: missing employee_id '{emp_id}' or name '{name}'; skipping"
                        )
                        continue
                    
                    # Get or create employee
                    emp, created_flag = Employee.objects.get_or_create(
                        employee_id=emp_id,
                        defaults={'name': name}
                    )
                    changed = False
                    
                    # Update name if different
                    if emp.name != name:
                        emp.name = name
                        changed = True
                    
                    # Update other fields if present
                    if mapping.get("email") is not None and len(row) > mapping["email"] and row[mapping["email"]]:
                        email = str(row[mapping["email"]]).strip()
                        if email and emp.email != email:
                            emp.email = email
                            changed = True
                    
                    if mapping.get("phone") is not None and len(row) > mapping["phone"] and row[mapping["phone"]]:
                        phone = str(row[mapping["phone"]]).strip()
                        if phone and emp.phone != phone:
                            emp.phone = phone
                            changed = True
                    
                    if mapping.get("department") is not None and len(row) > mapping["department"] and row[mapping["department"]]:
                        dept = str(row[mapping["department"]]).strip()
                        if dept and emp.department != dept:
                            emp.department = dept
                            changed = True
                    
                    if mapping.get("designation") is not None and len(row) > mapping["designation"] and row[mapping["designation"]]:
                        desig = str(row[mapping["designation"]]).strip()
                        if desig and emp.designation != desig:
                            emp.designation = desig
                            changed = True
                    
                    if mapping.get("monthly_salary") is not None and len(row) > mapping["monthly_salary"] and row[mapping["monthly_salary"]]:
                        try:
                            salary = float(str(row[mapping["monthly_salary"]]).strip())
                            if salary and emp.monthly_salary != salary:
                                emp.monthly_salary = salary
                                changed = True
                        except ValueError:
                            import_errors.append(f"Row {row_idx}: invalid salary value")
                    
                    if mapping.get("hire_date") is not None and len(row) > mapping["hire_date"] and row[mapping["hire_date"]]:
                        hire_date = parse_any_date(
                            str(row[mapping["hire_date"]]).strip(),
                            row[mapping["hire_date"]]
                        )
                        if hire_date and emp.hire_date != hire_date:
                            emp.hire_date = hire_date
                            changed = True
                    
                    # Import image if available
                    if mapping.get("employee_image_file") is not None and len(row) > mapping["employee_image_file"] and row[mapping["employee_image_file"]]:
                        img_path = str(row[mapping["employee_image_file"]]).strip()
                        if img_path:
                            full_img_path = os.path.join(temp_dir, img_path)
                            if os.path.exists(full_img_path):
                                with open(full_img_path, 'rb') as img_file:
                                    emp.employee_image.save(f"{emp_id}.jpg", File(img_file), save=False)
                                    changed = True
                    
                    if created_flag:
                        emp.save()
                        created += 1
                    elif changed:
                        emp.save()
                        updated += 1
                        
                except Exception as e:
                    import_errors.append(f"Row {row_idx}: unexpected error: {e}")
    
    return import_errors, {"created": created, "updated": updated}